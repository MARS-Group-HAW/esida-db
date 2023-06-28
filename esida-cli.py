import os
import json
import shutil
import pkgutil
import importlib
import subprocess
import datetime as dt
from pathlib import Path

import click
import shapely
import shapely.wkt
import geopandas
import pandas as pd
import sqlalchemy
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

import log
from dbconf import get_engine, connect, close, get_conn_string
from parameters import *
from esida.tiff_parameter import TiffParameter
from esida.models import Signal, Shape


logger = log.setup_custom_logger('root')

@click.group()
def cli():
    pass

@cli.command("load-shapes")
@click.argument('file', type=click.Path(exists=True))
def load_shapes(file):
    """ Load given shapefile structure into the database. """
    gdf = geopandas.read_file(file)
    required_cols = ['id', 'name', 'type', 'parent_id', 'properties', 'geometry']

    # optional cols, if not set make it empty
    if "properties" not in gdf.columns:
        gdf['properties'] = "{}"

    # make sure our required columns exist
    for col in required_cols:
        if col not in gdf.columns:
            raise click.UsageError(f"{col} column is required")

    # In Geopandas/Pandas there is no None/Null value for Integers. So our
    # parent_id column with the int reference is of type float (b/c upper most
    # shapes have no parent_id).
    # SQLAlchemy / Geopandas to_postgis() can't handle a float in enforcing the
    # foreign key to the parent row.
    # ...so we need to split our input data in parent only rows and child rows
    # and import them separately.
    gdf_no_parent = gdf[gdf['parent_id'].isna()].copy()
    gdf_no_parent[['id', 'name', 'type', 'properties', 'geometry']].to_postgis('shape', connect(), if_exists='append')

    gdf_with_parent = gdf[gdf['parent_id'].notna()].copy()
    gdf_with_parent['parent_id'] = gdf_with_parent['parent_id'].astype('int') # no isna() rows left => cast to int, so SQLAlchemy can write to PostGis
    gdf_with_parent[required_cols].to_postgis('shape', connect(), if_exists='append')

    # calculate shape area
    # Date are in ESPG:4326 (deg based), so for ST_Area() to produce m2
    # we need to convert to a meters based system. With utmzone() we identify
    # the resp. used UTM Zone that is m based.
    con = connect()
    con.execute('UPDATE shape SET area_sqm = \
        ST_Area(ST_Transform(geometry, utmzone(ST_Centroid(geometry))))')

@cli.command("dump")
@click.option('-o', '--output', type=click.Path())
def dump(output):
    """ Dump shapes and calculated Data Layers from the database using pg_dump. """

    if not output:
        output = f"{dt.datetime.today().strftime('%Y-%m-%d_%H-%M-%S')}_datahub.dump"

    try:
        params = ['pg_dump', "-Fc", "-f", f"output/{output}", get_conn_string()]
        subprocess.check_output(params)
    except subprocess.CalledProcessError as error:
        print(error.stderr)

@cli.command("restore")
@click.argument('file', type=click.Path(exists=True))
def restore(file):
    """ Restore shapes and calculated Data Layers from a previous dump. """

    click.confirm('This will overwrite previous data. Do you want to continue?', abort=True)

    # --clean:     drop all tables, before creating them
    # --if-exists: prevent DROP table -> table does not exist warnings during
    #              import in a empty database
    try:
        params = ['pg_restore', "--clean", "--if-exists", "-d", get_conn_string(), file]
        subprocess.check_output(params)
        return True
    except subprocess.CalledProcessError as error:
        print(error.stderr)

@cli.command("list")
def list_parameters():
    """ Print all available parameters. """
    parameters  = [name for _, name, _ in pkgutil.iter_modules(['parameters'])]

    for parameter in parameters:
        print(parameter)

@cli.command()
@click.argument('parameter')
@click.argument('action')
def param(parameter, action):
    """ Execute action on parameter, like load, transform or extract. """

    params  = [name for _, name, _ in pkgutil.iter_modules(['parameters'])]

    if parameter not in params:
        click.echo(click.style(f'Unknown parameter: {parameter}',  fg='red'), err=True)
        return

    pmodule = importlib.import_module(f'parameters.{parameter}')
    pclass  = getattr(pmodule, parameter)()
    result = getattr(pclass, action)()


@cli.command()
@click.argument('parameter', required=False, type=str)
@click.option('--shape_id', required=False, type=int)
def daspatial(parameter, shape_id):
    """ Generate long running spatial quality of applicable parameters. """

    if not parameter:
        params  = [name for _, name, _ in pkgutil.iter_modules(['parameters'])]
    else:
        params = [parameter]
    summary = []

    now = dt.datetime.now()
    out_name = now.strftime("%Y-%m-%d_%H-%M-%S")
    out_name += "_DA_Spatial"
    out_dir = Path("./output") / out_name
    out_dir.mkdir(parents=True, exist_ok=True)


    for p in params:
        pmodule = importlib.import_module(f'parameters.{p}')
        pc  = getattr(pmodule, p)()

        if not isinstance(pc, TiffParameter):
            continue

        result, results = pc.da_spatial(shape_id)

        summary.append({
            'data_layer': p,
            'spatial_coverage': result,
        })

        # save summary in each loop iteration to save output early.
        # generation takes long!
        dfx = pd.DataFrame(summary)
        dfx.to_csv(out_dir / 'da_spatial.csv', index=False)

        dfx = pd.DataFrame(results)
        dfx.to_csv(out_dir / f'files_{p}.csv', index=False)


@cli.command()
@click.option('--wkt')
@click.option('--abm', is_flag=True)
def clip(wkt, abm):
    """ Get metrics for the given AOI, provide either a WKT string or file.

    If the abm flag is present only data layers relevant for the ABM are
    exported and a config.json is generated.

    """
    params  = [name for _, name, _ in pkgutil.iter_modules(['parameters'])]

    now = dt.datetime.now()
    out_name = now.strftime("%Y-%m-%d_%H-%M-%S") + "_clip"
    if os.path.exists(wkt):
        out_name += "_"+os.path.basename(wkt)
        with open(wkt, 'r') as file:
            wkt = file.read()

    geometry = shapely.wkt.loads(wkt)
    if geometry.geom_type != 'Polygon':
        click.echo(click.style('Only POLYGON is allowed as input geometry.',  fg='red'), err=True)

    out_dir = Path("./output") / out_name
    out_resources = out_dir / "resources"
    out_resources.mkdir(parents=True, exist_ok=True)
    shape = {
        'geometry': geometry,
        'name': 'text',
        'id': 0,
    }

    for p in params:
        if abm and p not in [
                'worldpop_popc',
                'geofabrik_pois',
                'osm_graph',
                'osm_building',
                'osm_landuse',
                'rcmrd_elev',
                'meteo_tprecit',
                'visualcrossing_weather'
            ]:
            continue

        pmodule = importlib.import_module(f'parameters.{p}')
        pclass  = getattr(pmodule, p)()
        pclass.set_output_path(out_resources)
        pclass.output = 'fs' # save products to file system instead of database
        result = getattr(pclass, 'load')(shapes=[shape], save_output=True)

    # the following prepared the config.json for the MARS ABM
    # so if we don't have the abm flag this is not necessary.
    if not abm:
        return

    # after creating layers
    # calculate amount of needed agents
    dfx = pd.read_csv(out_resources / 'worldpop_popc.csv')
    agent_count = int(dfx['worldpop_popc'].tail(1))

    p = 'cia_worldfactbook'
    pmodule = importlib.import_module(f'parameters.{p}')
    pclass  = getattr(pmodule, p)()
    pclass.set_output_path(out_resources)
    pclass.output = 'fs' # save products to file system instead of database
    result = getattr(pclass, 'load')(shapes=[shape], save_output=True, agent_count=agent_count)

    shutil.copytree("input/data/MARS", out_dir.as_posix())

    # prepare config.json
    with open('input/data/MARS/config.json') as f:
        d = json.load(f)

        # set date
        start = dt.datetime.now()
        end   = start + dt.timedelta(days=14+1) # +1 so we don't need to run until 23:59:59
        d['globals']['startPoint'] = start.strftime('%Y-%m-%dT00:00:00')
        d['globals']['endPoint']   = end.strftime('%Y-%m-%dT00:00:00')

        # Geometries for temp/prcp CSV spatial join
        geometry_json = shapely.geometry.mapping(geometry)
        for d2 in d['layers']:
            if d2['name'] in ['TemperatureLayer', 'PrecipitationLayer']:
                for d3 in d2['inputs']:
                    if "value" in d3:
                        d3['value']['geometry'] = geometry_json

        # set agent count
        for d2 in d['agents']:
            if d2['name'] == 'Human':
                d2['count'] = agent_count

        with open(out_dir / 'config.json', 'w') as f:
            json.dump(d, f, indent=2)


@cli.command("signal")
@click.option('--id')
def assess_signal(id):
    signal = Signal.query.get(id)

    print("===")
    print(f"Signal-ID: {signal.id}")
    print(f"Signal date: {signal.report_date}")
    print("===")

    with open("input/algorithms/dengue-fever.yml", "r") as stream:
        try:
            algorithm = yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)

    results = {}

    # Process specs
    for shape in signal.shapes():
        spec = algorithm['spec'][algorithm['metadata']['start']]
        print(f"# {shape.name}")

        results[shape.id] = []

        while True:
            print(f" - Test: {spec['name']}")
            dl = spec['datalayer']
            pm = importlib.import_module(f'parameters.{dl}')
            pc = getattr(pm, dl)()


            r = {
                'spec': spec,
                'ops': []
            }

            bools = []

            for op in spec['operators']:
                b, df = shape.op(signal, pc, op['op'], op['attrs'])
                print(f"    └ {op['op']}: {b}")
                bools.append(b)
                r['ops'].append([b, df])

            positive = spec['positive']
            negative = spec['negative']
            if all(bools):
                results[shape.id].append(r)
                if positive and positive in algorithm['spec']:
                    spec = algorithm['spec'][positive]
                else:
                    results[shape.id].append({
                        'spec': {
                            'name': "Algorithm finished with POSITIVE outcome"
                        }
                    })
                    print("Algorithm finished with POSITIVE outcome")
                    break
            else:
                if negative and negative in algorithm['spec']:
                    spec = algorithm['spec'][negative]
                else:
                    results[shape.id].append({
                        'spec': {
                            'name': "Algorithm finished with NEGATIVE outcome"
                        }
                    })
                    print("Algorithm finished with NEGATIVE outcome")
                    break

        print("")


@cli.command("master-table")
@click.argument('file', type=click.Path(exists=True))
def master_table(file):

    with open(file, "r") as stream:
        try:
            algorithm = yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)
            return

    colors = {
        'country': Color(rgb='FFC7CE'),
        'region':  Color(rgb='FCE1C6'),
        'missing': Color(rgb='FF0000'),
    }

    wb = Workbook()
    ws = wb.active

    ws.title = "README"

    ws['A1'].value = "Inferred from country"
    ws['A1'].fill = PatternFill(fgColor=colors['country'], fill_type = "solid")
    ws['A2'].value = "Inferred from region"
    ws['A2'].fill = PatternFill(fgColor=colors['region'], fill_type = "solid")
    ws['A3'].value = "Missing value"
    ws['A3'].fill = PatternFill(fgColor=colors['missing'], fill_type = "solid")

    ws['A5'].value = f"Created at {dt.datetime.now().isoformat()}"

    for shape_type in ['country', 'region', 'district']:

        ws = wb.create_sheet(shape_type.capitalize())
        ws.freeze_panes = "A2" # freeze rows ABOVE A2 -> first row
        shapes = Shape.query.where(Shape.type == shape_type).order_by(Shape.name.asc()).all()

        col = 1
        row = 1

        # First row is Type
        ws.cell(row=row, column=col, value='Type')
        row += 1

        for s in shapes:
            d = ws.cell(row=row, column=col, value=s.type)
            row += 1
        row = 1
        col += 1

        # Second row is Name
        ws.cell(row=row, column=col, value='Name')
        row += 1

        for s in shapes:
            d = ws.cell(row=row, column=col, value=s.name)
            row += 1
        row = 1
        col += 1

        # Columns for parents
        if shape_type == 'district':
            ws.cell(row=row, column=col, value='Region')
            row += 1

            for s in shapes:
                d = ws.cell(row=row, column=col, value=s.parent.name)
                row += 1
            row = 1
            col += 1

        for dl in algorithm['spec']:
            pm = importlib.import_module(f"parameters.{dl['datalayer']}")
            pc = getattr(pm, dl['datalayer'])()

            d = ws.cell(row=row, column=col, value=dl['datalayer'])
            row += 1

            for s in shapes:
                v = s.get(pc, fallback_parent=True)



                if v and dl['datalayer'] in v:

                    value = v[dl['datalayer']]

                    # in case the data layer is a count, but we need a proportion
                    # load the corresponding total data layer
                    if 'datalayer_total' in dl:
                        total = s.get(dl['datalayer_total'])
                        total = total[dl['datalayer_total']]

                        value = value / total * 100


                    if pc.is_percent:
                        if not pc.is_percent100:
                            value = value * 100

                    if f"{dl['datalayer']}_inferred" in v:
                        c = ws.cell(row=row, column=col, value=value)
                        c.fill = PatternFill(fgColor=colors[v['type']], fill_type = "solid")
                    else:
                        ws.cell(row=row, column=col, value=value)
                else:
                    c = ws.cell(row=row, column=col, value="=NA()")

                row += 1

            row= 1
            col += 1

        # auto width for all columns
        if False:
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value


    wb.save('DataHub_MASTER.xlsx')




if __name__ == '__main__':
    cli()

    # make sure potentially open database connections are closed
    close()
